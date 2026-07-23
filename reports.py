import sqlite3
from datetime import datetime
from PIL import Image, ImageDraw, ImageFont
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

THEMES = {
    "light": {
        "bg_top": (255, 223, 240),
        "bg_bottom": (200, 225, 255),
        "card": (255, 255, 255, 235),
        "accent": (60, 110, 220),
        "text_dark": (35, 35, 45),
        "text_gray": (120, 120, 135),
        "line_colors": ["#E8467D", "#4C7DF0", "#F2B33D", "#2FBF9F", "#9B6DE0"],
        "pie_colors": ["#E8467D", "#4C7DF0", "#F2B33D", "#2FBF9F", "#9B6DE0", "#2B2B36"],
        "table_header": (66, 118, 224),
        "row_alt": (245, 248, 255),
    },
    "dark": {
        "bg_top": (28, 28, 42),
        "bg_bottom": (12, 12, 22),
        "card": (42, 42, 58, 235),
        "accent": (140, 160, 255),
        "text_dark": (235, 235, 240),
        "text_gray": (175, 175, 185),
        "line_colors": ["#FF6B9D", "#6FA8FF", "#FFD166", "#5FE0BE", "#C79BFF"],
        "pie_colors": ["#FF6B9D", "#6FA8FF", "#FFD166", "#5FE0BE", "#C79BFF", "#8A8FA0"],
        "table_header": (70, 90, 170),
        "row_alt": (36, 36, 50),
    },
}


def get_connection():
    conn = sqlite3.connect("attendance.db")
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def get_setting(key, default=None):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT)")
    cursor.execute("SELECT value FROM settings WHERE key = ?", (key,))
    row = cursor.fetchone()
    conn.commit()
    conn.close()
    return row[0] if row else default


def set_setting(key, value):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT)")
    cursor.execute(
        "INSERT INTO settings (key, value) VALUES (?, ?) "
        "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
        (key, value)
    )
    conn.commit()
    conn.close()


def get_student_theme(student_id):
    conn = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT theme FROM students WHERE student_id = ?", (student_id,))
        row = cursor.fetchone()
    except Exception:
        row = None
    conn.close()
    if row and row[0] in THEMES:
        return row[0]
    return "light"


def set_student_theme(student_id, theme):
    if theme not in THEMES:
        return False
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE students SET theme = ? WHERE student_id = ?", (theme, student_id))
    conn.commit()
    conn.close()
    return True


def get_student_data(student_id):
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT name, roll_number FROM students WHERE student_id = ?", (student_id,))
    student_row = cursor.fetchone()
    if not student_row:
        conn.close()
        return None
    name, roll_number = student_row

    cursor.execute(
        """SELECT c.course_id, c.course_name, c.section
           FROM courses c
           JOIN enrollments e ON c.course_id = e.course_id
           WHERE e.student_id = ?""",
        (student_id,)
    )
    courses = cursor.fetchall()

    course_stats = []
    course_series = {}
    total_all = 0
    present_all = 0
    all_dates = []

    for course_id, course_name, section in courses:
        cursor.execute(
            """SELECT s.session_date, ar.status FROM attendance_records ar
               JOIN attendance_sessions s ON ar.session_id = s.session_id
               WHERE ar.student_id = ? AND s.course_id = ?
               ORDER BY s.session_date, s.start_time""",
            (student_id, course_id)
        )
        records = cursor.fetchall()
        total = len(records)
        present = sum(1 for r in records if r[1] == "present")
        pct = (present / total * 100) if total > 0 else 0.0
        if_miss = (present / (total + 1) * 100) if total >= 0 else 0.0
        if_attend = ((present + 1) / (total + 1) * 100) if total >= 0 else 0.0

        course_stats.append({
            "code": f"{course_name} ({section})",
            "total": total,
            "present": present,
            "absent": total - present,
            "pct": pct,
            "if_miss": if_miss,
            "if_attend": if_attend,
        })

        total_all += total
        present_all += present

        series = []
        running_present = 0
        for i, (date_str, status) in enumerate(records, start=1):
            if status == "present":
                running_present += 1
            series.append((date_str, running_present / i * 100))
            all_dates.append(date_str)
        course_series[f"{course_name}"] = series

    cursor.execute(
        """SELECT s.session_date, ar.status FROM attendance_records ar
           JOIN attendance_sessions s ON ar.session_id = s.session_id
           WHERE ar.student_id = ?
           ORDER BY s.session_date ASC, s.start_time ASC""",
        (student_id,)
    )
    all_records_asc = [r[1] for r in cursor.fetchall()]

    current_streak = 0
    for status in reversed(all_records_asc):
        if status == "present":
            current_streak += 1
        else:
            break

    highest_streak = 0
    running = 0
    for status in all_records_asc:
        if status == "present":
            running += 1
            highest_streak = max(highest_streak, running)
        else:
            running = 0

    conn.close()

    overall_pct = (present_all / total_all * 100) if total_all > 0 else 0.0
    date_range = ""
    if all_dates:
        date_range = f"{min(all_dates)} - {max(all_dates)}"

    return {
        "name": name,
        "roll_number": roll_number,
        "total_classes": total_all,
        "present_classes": present_all,
        "overall_pct": overall_pct,
        "courses": course_stats,
        "course_series": course_series,
        "streak": current_streak,
        "highest_streak": highest_streak,
        "date_range": date_range,
    }


def _load_font(size, bold=False):
    candidates = [
        "arialbd.ttf" if bold else "arial.ttf",
        "DejaVuSans-Bold.ttf" if bold else "DejaVuSans.ttf",
    ]
    for name in candidates:
        try:
            return ImageFont.truetype(name, size)
        except Exception:
            continue
    return ImageFont.load_default()


def _make_gradient(width, height, top_color, bottom_color):
    base = Image.new("RGB", (width, height), top_color)
    top = Image.new("RGB", (width, height), bottom_color)
    mask = Image.new("L", (width, height))
    mask_data = []
    for y in range(height):
        mask_data.extend([int(255 * (y / height))] * width)
    mask.putdata(mask_data)
    base.paste(top, (0, 0), mask)
    return base


def _draw_rounded_translucent(base_img, box, radius, fill_rgba):
    overlay = Image.new("RGBA", base_img.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay)
    draw.rounded_rectangle(box, radius=radius, fill=fill_rgba)
    base_img.paste(Image.alpha_composite(base_img.convert("RGBA"), overlay).convert("RGB"), (0, 0))


def _make_line_chart(course_series, line_colors, theme_key, out_path):
    fig, ax = plt.subplots(figsize=(5.7, 3.0), dpi=150)
    fig.patch.set_alpha(0)
    ax.set_facecolor("none")
    text_color = "#DDDDDD" if theme_key == "dark" else "#555555"

    has_data = False
    n_courses = 0
    for i, (course_name, series) in enumerate(course_series.items()):
        if not series:
            continue
        has_data = True
        n_courses += 1
        dates = [s[0] for s in series]
        pcts = [s[1] for s in series]
        color = line_colors[i % len(line_colors)]
        ax.plot(dates, pcts, label=course_name, color=color, linewidth=2.2)

    ax.set_ylim(0, 105)
    ax.grid(True, alpha=0.15)
    ax.tick_params(colors=text_color, labelsize=6.5)
    for spine in ax.spines.values():
        spine.set_color(text_color)
        spine.set_alpha(0.25)
    if has_data:
        ax.legend(
            fontsize=7, labelcolor=text_color, frameon=False,
            loc="lower center", bbox_to_anchor=(0.5, 1.02),
            ncol=min(n_courses, 4), handlelength=1.2, columnspacing=1.0
        )
        plt.xticks(rotation=35, ha="right")
    fig.tight_layout()
    fig.savefig(out_path, transparent=True)
    plt.close(fig)


def _make_pie_chart(courses, pie_colors, theme_key, out_path):
    labels = [c["code"].split(" (")[0] for c in courses]
    values = [c["present"] for c in courses]
    total_absent = sum(c["absent"] for c in courses)
    labels.append("Absent")
    values.append(total_absent)

    text_color = "#DDDDDD" if theme_key == "dark" else "#555555"

    fig, ax = plt.subplots(figsize=(3.4, 3.0), dpi=150)
    fig.patch.set_alpha(0)
    if sum(values) > 0:
        wedges, _, _ = ax.pie(
            values, colors=pie_colors[:len(values)],
            autopct="%1.0f%%", textprops={"fontsize": 7, "color": "white"},
            wedgeprops={"linewidth": 1, "edgecolor": "white" if theme_key == "light" else "#1c1c2a"},
            pctdistance=0.72
        )
        ax.legend(
            wedges, labels, fontsize=6.5, labelcolor=text_color, frameon=False,
            loc="lower center", bbox_to_anchor=(0.5, 1.0), ncol=3, handlelength=1.0
        )
    fig.tight_layout()
    fig.savefig(out_path, transparent=True)
    plt.close(fig)


def generate_report_image(student_id, output_path, theme="light"):
    data = get_student_data(student_id)
    if not data:
        return None

    if theme not in THEMES:
        theme = "light"
    t = THEMES[theme]

    institute_name = get_setting("institute_name", "Attendance Bot")

    width = 1000
    n_courses = len(data["courses"])
    table_row_h = 42
    table_h = table_row_h * (n_courses + 1)
    height = 970 + (table_h - 200 if table_h > 200 else 0)
    height = max(height, 950)

    img = _make_gradient(width, height, t["bg_top"], t["bg_bottom"])
    draw = ImageDraw.Draw(img)

    font_title = _load_font(24, bold=True)
    font_sub = _load_font(14)
    font_name = _load_font(16, bold=True)
    font_roll = _load_font(13)
    font_stat_num = _load_font(28, bold=True)
    font_stat_label = _load_font(13)
    font_stat_sub = _load_font(10)
    font_table_header = _load_font(13, bold=True)
    font_table_cell = _load_font(12)
    font_footer = _load_font(12)

    header_h = 95
    _draw_rounded_translucent(img, [30, 25, width - 30, 25 + header_h], 18, t["card"])
    draw = ImageDraw.Draw(img)

    logo_cx, logo_cy, logo_r = 70, 25 + header_h // 2, 28
    draw.ellipse([logo_cx - logo_r, logo_cy - logo_r, logo_cx + logo_r, logo_cy + logo_r],
                 outline=t["accent"], width=3)
    initials = "".join([w[0] for w in institute_name.split()[:2]]).upper() or "AB"
    draw.text((logo_cx, logo_cy), initials, font=font_stat_sub, fill=t["accent"], anchor="mm")

    title_display = institute_name if len(institute_name) <= 42 else institute_name[:39] + "..."
    draw.text((115, 25 + 14), title_display, font=font_title, fill=t["accent"])
    draw.text((115, 25 + 52), "Attendance Bot  ·  Student Report", font=font_sub, fill=t["text_gray"])

    name_text = data["name"]
    roll_text = f"Roll No: {data['roll_number']}"
    name_w = draw.textlength(name_text, font=font_name)
    roll_w = draw.textlength(roll_text, font=font_roll)
    draw.text((width - 50 - name_w, 25 + 22), name_text, font=font_name, fill=t["text_dark"])
    draw.text((width - 50 - roll_w, 25 + 48), roll_text, font=font_roll, fill=t["text_gray"])

    stats_top = 140
    stats_h = 130
    gap = 22
    card_w = (width - 60 - gap * 3) // 4
    stats = [
        ("Total Classes", f"{data['present_classes']}/{data['total_classes']}", data["date_range"]),
        ("Attendance %", f"{data['overall_pct']:.1f}%", data["date_range"]),
        ("Current Streak", f"{data['streak']} days", f"Highest {data['highest_streak']} days"),
        ("Courses Enrolled", f"{n_courses}", "Active this term"),
    ]
    for i, (label, value, sub) in enumerate(stats):
        x = 30 + i * (card_w + gap)
        _draw_rounded_translucent(img, [x, stats_top, x + card_w, stats_top + stats_h], 16, t["card"])
    draw = ImageDraw.Draw(img)
    for i, (label, value, sub) in enumerate(stats):
        x = 30 + i * (card_w + gap)
        draw.text((x + 18, stats_top + 18), label, font=font_stat_label, fill=t["text_gray"])
        draw.text((x + 18, stats_top + 42), value, font=font_stat_num, fill=t["accent"])
        draw.text((x + 18, stats_top + 90), sub, font=font_stat_sub, fill=t["text_gray"])

    table_top = stats_top + stats_h + 30
    table_x = 30
    table_w = width - 60
    headers = ["Class Code", "Total", "Present", "Present %", "If Miss", "If Attend"]
    col_ratios = [0.32, 0.11, 0.13, 0.15, 0.14, 0.15]
    col_widths = [int(table_w * r) for r in col_ratios]

    _draw_rounded_translucent(img, [table_x, table_top, table_x + table_w, table_top + table_h], 16, t["card"])
    draw = ImageDraw.Draw(img)

    header_overlay = Image.new("RGBA", img.size, (0, 0, 0, 0))
    hdraw = ImageDraw.Draw(header_overlay)
    hdraw.rounded_rectangle(
        [table_x, table_top, table_x + table_w, table_top + table_row_h],
        radius=16, fill=t["table_header"] + (255,)
    )
    hdraw.rectangle([table_x, table_top + 16, table_x + table_w, table_top + table_row_h], fill=t["table_header"] + (255,))
    img.paste(Image.alpha_composite(img.convert("RGBA"), header_overlay).convert("RGB"), (0, 0))
    draw = ImageDraw.Draw(img)

    x = table_x
    for h, w in zip(headers, col_widths):
        draw.text((x + 14, table_top + 12), h, font=font_table_header, fill=(255, 255, 255))
        x += w

    y = table_top + table_row_h
    for idx, course in enumerate(data["courses"]):
        if idx % 2 == 1:
            draw.rectangle([table_x, y, table_x + table_w, y + table_row_h], fill=t["row_alt"])
        x = table_x
        row_vals = [
            course["code"], str(course["total"]), str(course["present"]),
            f"{course['pct']:.1f}%", f"{course['if_miss']:.1f}%", f"{course['if_attend']:.1f}%"
        ]
        for val, w in zip(row_vals, col_widths):
            draw.text((x + 14, y + 12), val, font=font_table_cell, fill=t["text_dark"])
            x += w
        y += table_row_h

    charts_top = table_top + table_h + 30
    charts_h = 330
    _draw_rounded_translucent(img, [30, charts_top, width - 30, charts_top + charts_h], 18, t["card"])
    draw = ImageDraw.Draw(img)

    line_path = output_path + "_line.png"
    pie_path = output_path + "_pie.png"
    _make_line_chart(data["course_series"], t["line_colors"], theme, line_path)
    _make_pie_chart(data["courses"], t["pie_colors"], theme, pie_path)

    try:
        line_img = Image.open(line_path).convert("RGBA")
        line_img = line_img.resize((580, charts_h - 30))
        img.paste(line_img, (50, charts_top + 15), line_img)
    except Exception:
        pass

    try:
        pie_img = Image.open(pie_path).convert("RGBA")
        pie_img = pie_img.resize((320, charts_h - 30))
        img.paste(pie_img, (650, charts_top + 15), pie_img)
    except Exception:
        pass

    import os
    for p in (line_path, pie_path):
        try:
            os.remove(p)
        except Exception:
            pass

    footer_top = charts_top + charts_h + 20
    draw = ImageDraw.Draw(img)
    pill_text = "Generated by Attendance Bot"
    pill_w = draw.textlength(pill_text, font=font_footer) + 40
    _draw_rounded_translucent(img, [30, footer_top, 30 + pill_w, footer_top + 36], 18, t["card"])
    draw = ImageDraw.Draw(img)
    draw.text((30 + 20, footer_top + 9), pill_text, font=font_footer, fill=t["accent"])
    draw.text((30, footer_top + 45), datetime.now().strftime("%d %b %Y, %H:%M"), font=font_stat_sub, fill=t["text_gray"])

    final_height = footer_top + 80
    img = img.crop((0, 0, width, final_height))

    img.save(output_path)
    return output_path


def get_date_grid_data(student_id):
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT name, roll_number FROM students WHERE student_id = ?", (student_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return None
    name, roll_number = row

    cursor.execute(
        """SELECT c.course_id, c.course_name, c.section
           FROM courses c
           JOIN enrollments e ON c.course_id = e.course_id
           WHERE e.student_id = ?""",
        (student_id,)
    )
    courses = cursor.fetchall()

    course_codes = [f"{cn} ({sec})" for _, cn, sec in courses]
    grid = {code: {} for code in course_codes}
    all_dates = set()
    course_totals = {}

    for cid, cn, sec in courses:
        code = f"{cn} ({sec})"
        cursor.execute(
            """SELECT s.session_date, ar.status FROM attendance_records ar
               JOIN attendance_sessions s ON ar.session_id = s.session_id
               WHERE ar.student_id = ? AND s.course_id = ?
               ORDER BY s.session_date, s.session_id""",
            (student_id, cid)
        )
        recs = cursor.fetchall()
        day_status = {}
        for d, st in recs:
            day_status[d] = st  # latest session for that date wins

        total = 0
        present = 0
        for d, st in day_status.items():
            grid[code][d] = "P" if st == "present" else "A"
            all_dates.add(d)
            total += 1
            if st == "present":
                present += 1

        pct = (present / total * 100) if total > 0 else 0.0
        course_totals[code] = {
            "total": total, "present": present,
            "absent": total - present, "pct": pct
        }

    conn.close()
    sorted_dates = sorted(all_dates)

    return {
        "name": name,
        "roll_number": roll_number,
        "course_codes": course_codes,
        "dates": sorted_dates,
        "grid": grid,
        "course_totals": course_totals,
    }


def generate_report_pdf(student_id, output_path):
    grid_data = get_date_grid_data(student_id)
    if not grid_data:
        return None

    institute_name = get_setting("institute_name", "Attendance Bot")
    name = grid_data["name"]
    roll_number = grid_data["roll_number"]
    dates = grid_data["dates"]
    codes = grid_data["course_codes"]
    grid = grid_data["grid"]
    totals = grid_data["course_totals"]

    def fmt_date(d):
        try:
            dt = datetime.strptime(d, "%Y-%m-%d")
            return dt.strftime("%d-%b-%Y")
        except Exception:
            return d

    DATES_PER_PAGE = 11
    if dates:
        chunks = [dates[i:i + DATES_PER_PAGE] for i in range(0, len(dates), DATES_PER_PAGE)]
    else:
        chunks = [[]]
    total_pages = len(chunks)

    page_w, page_h = landscape(A4)

    doc = SimpleDocTemplate(
        output_path, pagesize=landscape(A4),
        topMargin=22 * mm, bottomMargin=18 * mm,
        leftMargin=10 * mm, rightMargin=10 * mm
    )

    elements = []

    for page_idx, chunk in enumerate(chunks):
        is_last = (page_idx == total_pages - 1)

        header_row = ["Class Code"] + [fmt_date(d) for d in chunk]
        if is_last:
            header_row += ["Total", "Present", "Present(%)", "Absent", "Absent(%)"]

        table_rows = [header_row]
        for code in codes:
            row = [code]
            for d in chunk:
                row.append(grid[code].get(d, ""))
            if is_last:
                ct = totals[code]
                absent_pct = 100 - ct["pct"] if ct["total"] > 0 else 0
                row += [
                    str(ct["total"]), str(ct["present"]),
                    f"{ct['pct']:.2f} %", str(ct["absent"]), f"{absent_pct:.2f} %"
                ]
            table_rows.append(row)

        total_classes_row = ["Total Classes"]
        total_present_row = ["Total Present Classes"]
        present_pct_row = ["Present %"]
        total_absent_row = ["Total Absent Classes"]
        absent_pct_row = ["Absent %"]

        for d in chunk:
            day_total = sum(1 for code in codes if d in grid[code])
            day_present = sum(1 for code in codes if grid[code].get(d) == "P")
            day_absent = day_total - day_present
            day_pct = (day_present / day_total * 100) if day_total > 0 else 0
            day_absent_pct = 100 - day_pct if day_total > 0 else 0

            total_classes_row.append(str(day_total))
            total_present_row.append(str(day_present))
            present_pct_row.append(f"{day_pct:.0f} %")
            total_absent_row.append(str(day_absent))
            absent_pct_row.append(f"{day_absent_pct:.0f} %")

        if is_last:
            for r in (total_classes_row, total_present_row, present_pct_row,
                      total_absent_row, absent_pct_row):
                r += ["", "", "", "", ""]

        table_rows += [total_classes_row, total_present_row, present_pct_row,
                        total_absent_row, absent_pct_row]

        n_cols = len(header_row)
        available_w = page_w - 20 * mm
        col_width = available_w / n_cols

        t = Table(table_rows, colWidths=[col_width] * n_cols, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4276E0")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 6.5),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0, 1), (-1, len(codes)), [colors.white, colors.HexColor("#F5F8FF")]),
            ("BACKGROUND", (0, len(codes) + 1), (-1, -1), colors.HexColor("#EFF3FF")),
        ]))
        elements.append(t)

        if page_idx < total_pages - 1:
            elements.append(PageBreak())

    def on_page(canvas, doc_):
        canvas.saveState()
        canvas.setFont("Helvetica-Bold", 12)
        canvas.drawString(15 * mm, page_h - 14 * mm, institute_name)
        canvas.setFont("Helvetica", 9)
        canvas.drawString(15 * mm, page_h - 20 * mm, f"{name}  |  Roll No: {roll_number}")
        canvas.setFont("Helvetica", 8)
        canvas.drawCentredString(page_w / 2, 10 * mm, f"Attendance Report of {name}")
        canvas.drawRightString(page_w - 15 * mm, 10 * mm, f"Page {canvas.getPageNumber()} of {total_pages}")
        canvas.restoreState()

    doc.build(elements, onFirstPage=on_page, onLaterPages=on_page)
    return output_path

def get_course_grid_data(course_id):
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT course_name, section FROM courses WHERE course_id = ?", (course_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return None
    course_name, section = row

    cursor.execute(
        """SELECT s.student_id, s.roll_number, s.name
           FROM students s
           JOIN enrollments e ON s.student_id = e.student_id
           WHERE e.course_id = ?
           ORDER BY s.roll_number""",
        (course_id,)
    )
    students = cursor.fetchall()

    all_dates = set()
    grid = {}
    totals = {}

    for sid, roll, name in students:
        cursor.execute(
            """SELECT s.session_date, ar.status FROM attendance_records ar
               JOIN attendance_sessions s ON ar.session_id = s.session_id
               WHERE ar.student_id = ? AND s.course_id = ?
               ORDER BY s.session_date, s.session_id""",
            (sid, course_id)
        )
        recs = cursor.fetchall()
        day_status = {}
        for d, st in recs:
            day_status[d] = st  # latest session for that date wins

        g = {}
        total = 0
        present = 0
        for d, st in day_status.items():
            g[d] = "P" if st == "present" else "A"
            all_dates.add(d)
            total += 1
            if st == "present":
                present += 1

        grid[sid] = g
        pct = (present / total * 100) if total > 0 else 0.0
        totals[sid] = {
            "roll": roll, "name": name, "total": total,
            "present": present, "absent": total - present, "pct": pct
        }

    conn.close()

    return {
        "course_name": course_name,
        "section": section,
        "students": students,
        "dates": sorted(all_dates),
        "grid": grid,
        "totals": totals,
    }


def generate_course_report_pdf(course_id, output_path):
    cdata = get_course_grid_data(course_id)
    if not cdata:
        return None

    institute_name = get_setting("institute_name", "Attendance Bot")
    course_name = cdata["course_name"]
    section = cdata["section"]
    students = cdata["students"]
    dates = cdata["dates"]
    grid = cdata["grid"]
    totals = cdata["totals"]

    def fmt_date(d):
        try:
            dt = datetime.strptime(d, "%Y-%m-%d")
            return dt.strftime("%d-%b-%Y")
        except Exception:
            return d

    DATES_PER_PAGE = 9
    if dates:
        chunks = [dates[i:i + DATES_PER_PAGE] for i in range(0, len(dates), DATES_PER_PAGE)]
    else:
        chunks = [[]]
    total_pages = len(chunks)

    page_w, page_h = landscape(A4)

    doc = SimpleDocTemplate(
        output_path, pagesize=landscape(A4),
        topMargin=22 * mm, bottomMargin=18 * mm,
        leftMargin=10 * mm, rightMargin=10 * mm
    )

    elements = []

    for page_idx, chunk in enumerate(chunks):
        is_last = (page_idx == total_pages - 1)

        header_row = ["Roll No", "Name"] + [fmt_date(d) for d in chunk]
        if is_last:
            header_row += ["Total", "Present", "Present(%)", "Absent"]

        table_rows = [header_row]
        for sid, roll, name in students:
            row = [roll, name]
            for d in chunk:
                row.append(grid.get(sid, {}).get(d, ""))
            if is_last:
                ct = totals[sid]
                row += [str(ct["total"]), str(ct["present"]), f"{ct['pct']:.1f} %", str(ct["absent"])]
            table_rows.append(row)

        daily_present_row = ["", "Present Count"]
        daily_total_row = ["", "Total Marked"]
        for d in chunk:
            day_total = sum(1 for sid, _, _ in students if d in grid.get(sid, {}))
            day_present = sum(1 for sid, _, _ in students if grid.get(sid, {}).get(d) == "P")
            daily_total_row.append(str(day_total))
            daily_present_row.append(str(day_present))
        if is_last:
            daily_total_row += ["", "", "", ""]
            daily_present_row += ["", "", "", ""]
        table_rows += [daily_total_row, daily_present_row]

        n_cols = len(header_row)
        available_w = page_w - 20 * mm
        name_w = 45 * mm
        roll_w = 25 * mm
        remaining = available_w - name_w - roll_w
        other_cols = n_cols - 2
        other_w = remaining / other_cols if other_cols > 0 else remaining
        col_widths = [roll_w, name_w] + [other_w] * other_cols

        t = Table(table_rows, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4276E0")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (2, 0), (-1, -1), "CENTER"),
            ("ALIGN", (0, 0), (1, -1), "LEFT"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("ROWBACKGROUNDS", (0, 1), (-1, len(students)), [colors.white, colors.HexColor("#F5F8FF")]),
            ("BACKGROUND", (0, len(students) + 1), (-1, -1), colors.HexColor("#EFF3FF")),
            ("FONTNAME", (0, len(students) + 1), (-1, -1), "Helvetica-Bold"),
        ]))
        elements.append(t)

        if page_idx < total_pages - 1:
            elements.append(PageBreak())

    def on_page(canvas, doc_):
        canvas.saveState()
        canvas.setFont("Helvetica-Bold", 12)
        canvas.drawString(15 * mm, page_h - 14 * mm, institute_name)
        canvas.setFont("Helvetica", 9)
        canvas.drawString(15 * mm, page_h - 20 * mm, f"{course_name} ({section}) - Attendance Report")
        canvas.setFont("Helvetica", 8)
        canvas.drawCentredString(page_w / 2, 10 * mm, f"Course Report - {course_name} ({section})")
        canvas.drawRightString(page_w - 15 * mm, 10 * mm, f"Page {canvas.getPageNumber()} of {total_pages}")
        canvas.restoreState()

    doc.build(elements, onFirstPage=on_page, onLaterPages=on_page)
    return output_path


def generate_course_report_excel(course_id, output_path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    cdata = get_course_grid_data(course_id)
    if not cdata:
        return None

    course_name = cdata["course_name"]
    section = cdata["section"]
    students = cdata["students"]
    dates = cdata["dates"]
    grid = cdata["grid"]
    totals = cdata["totals"]

    def fmt_date(d):
        try:
            dt = datetime.strptime(d, "%Y-%m-%d")
            return dt.strftime("%d-%b-%Y")
        except Exception:
            return d

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{course_name}_{section}"[:31]

    header_fill = PatternFill(start_color="4276E0", end_color="4276E0", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    alt_fill = PatternFill(start_color="F5F8FF", end_color="F5F8FF", fill_type="solid")

    headers = ["Roll No", "Name"] + [fmt_date(d) for d in dates] + ["Total", "Present", "Absent", "Present %"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for idx, (sid, roll, name) in enumerate(students, start=2):
        ct = totals[sid]
        row = [roll, name] + [grid.get(sid, {}).get(d, "") for d in dates] + [
            ct["total"], ct["present"], ct["absent"], round(ct["pct"], 2)
        ]
        ws.append(row)
        if idx % 2 == 0:
            for cell in ws[idx]:
                cell.fill = alt_fill

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    for i in range(3, 3 + len(dates)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 11

    ws.freeze_panes = "C2"

    wb.save(output_path)
    return output_path
